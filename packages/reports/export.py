"""Export: PDF (ReportLab), Excel (openpyxl), CSV. All include data_quality + run_id."""

from __future__ import annotations

import csv, hashlib, io, os, uuid
from datetime import datetime, timezone
from packages.domain.constants import (
    EXPORT_CSV_PATH,
    EXPORT_EXCEL_PATH,
    EXPORT_PDF_PATH,
)
from packages.domain.entities import ExportQuality, ExportResult

SYSTEM_VERSION = "0.9.1"
DEFAULT_SEED = 42
DISCLAIMER = (
    "Dieses Ergebnis dient ausschließlich der Analyse, Forschung und Bildung. "
    "Es stellt keine Finanzberatung und keine Kauf- oder Verkaufsempfehlung dar. "
    "Historische oder simulierte Ergebnisse sind keine Garantie für zukünftige "
    "Ergebnisse."
)


def report_metadata(
    run_id: str, dq: ExportQuality | None = None, seed: int | None = None
) -> dict:
    """Structured metadata block included in every export."""
    return {
        "system_version": SYSTEM_VERSION,
        "run_id": run_id,
        "created_at": _utc(),
        "seed": seed if seed is not None else DEFAULT_SEED,
        "data_quality": dq.to_dict() if dq else {},
        "disclaimer": DISCLAIMER,
    }


def _utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _run_id() -> str:
    return str(uuid.uuid4())[:12]


def _data_hash(data: bytes | str) -> str:
    if isinstance(data, str):
        data = data.encode()
    return hashlib.sha256(data).hexdigest()[:16]


def _to_csv(rows: list[dict]) -> str:
    if not rows:
        return ""
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    w.writeheader()
    w.writerows(rows)
    return buf.getvalue()


def _ensure_dir(path: str) -> str:
    os.makedirs(path, exist_ok=True)
    return path


def csv_trades(
    trades: list[dict], dq: ExportQuality | None = None, seed: int | None = None
) -> ExportResult:
    """Export trade log as CSV with metadata."""
    run_id = _run_id()
    csv_data = _to_csv(trades)
    dq = dq or ExportQuality(len(trades), 0.0, "unknown")
    fp = os.path.join(_ensure_dir(EXPORT_CSV_PATH), f"trades_{run_id}.csv")
    with open(fp, "w", newline="", encoding="utf-8") as f:
        f.write(csv_data)
    return ExportResult(
        run_id=run_id,
        format="csv",
        data_quality=dq,
        data_hash=_data_hash(csv_data),
        file_path=fp,
        file_bytes=csv_data.encode(),
        created_at=_utc(),
        metadata={
            **report_metadata(run_id, dq, seed),
            "kind": "trades",
            "rows": len(trades),
        },
    )


def csv_equity(
    curve: list[float],
    dates: list[str] | None = None,
    dq: ExportQuality | None = None,
    seed: int | None = None,
) -> ExportResult:
    """Export equity curve as CSV with metadata."""
    run_id = _run_id()
    rows = [
        {"step": i, "date": dates[i] if dates and i < len(dates) else "", "equity": v}
        for i, v in enumerate(curve)
    ]
    csv_data = _to_csv(rows)
    dq = dq or ExportQuality(len(curve), 0.0, "unknown")
    fp = os.path.join(_ensure_dir(EXPORT_CSV_PATH), f"equity_{run_id}.csv")
    with open(fp, "w", newline="", encoding="utf-8") as f:
        # metadata header line (documented: '# ' prefix, UTF-8, comma-delimited)
        meta_line = ",".join(
            f"# {k}: {v}"
            for k, v in report_metadata(run_id, dq, seed).items()
            if k != "data_quality"
        )
        f.write(meta_line + "\n")
        f.write(csv_data)
    return ExportResult(
        run_id=run_id,
        format="csv",
        data_quality=dq,
        data_hash=_data_hash(csv_data),
        file_path=fp,
        file_bytes=(meta_line + "\n" + csv_data).encode(),
        created_at=_utc(),
        metadata={
            **report_metadata(run_id, dq, seed),
            "kind": "equity",
            "rows": len(curve),
        },
    )


def csv_scenario(
    runs: list[dict], dq: ExportQuality | None = None, seed: int | None = None
) -> ExportResult:
    """Export scenario runs as CSV with metadata."""
    run_id = _run_id()
    csv_data = _to_csv(runs)
    dq = dq or ExportQuality(len(runs), 0.0, "unknown")
    fp = os.path.join(_ensure_dir(EXPORT_CSV_PATH), f"scenario_{run_id}.csv")
    with open(fp, "w", newline="", encoding="utf-8") as f:
        meta_line = ",".join(
            f"# {k}: {v}"
            for k, v in report_metadata(run_id, dq, seed).items()
            if k != "data_quality"
        )
        f.write(meta_line + "\n")
        f.write(csv_data)
    return ExportResult(
        run_id=run_id,
        format="csv",
        data_quality=dq,
        data_hash=_data_hash(csv_data),
        file_path=fp,
        file_bytes=(meta_line + "\n" + csv_data).encode(),
        created_at=_utc(),
        metadata={
            **report_metadata(run_id, dq, seed),
            "kind": "scenario",
            "rows": len(runs),
        },
    )


def pdf_report(
    title: str,
    metrics: dict,
    trades: list[dict],
    chart_images: list | None = None,
    equity_curve: list[float] | None = None,
    dq: ExportQuality | None = None,
    seed: int | None = None,
) -> ExportResult:
    """Build a PDF report as ExportResult. Charts are PIL Images."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Image,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    run_id = _run_id()
    meta = report_metadata(run_id, dq, seed)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    ts = ParagraphStyle("t", parent=styles["Title"], fontSize=18)
    ds = ParagraphStyle(
        "d", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#555555")
    )
    elements: list = [
        Paragraph(title, ts),
        Paragraph(
            f"Run: {run_id} · {meta['created_at']} · "
            f"Version: {meta['system_version']} · Seed: {meta['seed']}",
            styles["Normal"],
        ),
        Spacer(1, 3 * mm),
        Paragraph(meta["disclaimer"], ds),
        Spacer(1, 6 * mm),
    ]
    if metrics:
        elements.append(Paragraph("Metrics", styles["Heading2"]))
        mdata = [[k, f"{v}"] for k, v in metrics.items()]
        mt = Table(mdata, colWidths=[120, 200])
        mt.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f8f9fa")],
                    ),
                ]
            )
        )
        elements.extend([mt, Spacer(1, 6 * mm)])
    if chart_images:
        elements.append(Paragraph("Charts", styles["Heading2"]))
        for img in chart_images:
            try:
                bio = io.BytesIO()
                img.save(bio, format="PNG")
                bio.seek(0)
                elements.extend(
                    [Image(bio, width=160 * mm, height=90 * mm), Spacer(1, 4 * mm)]
                )
            except Exception:
                elements.append(Paragraph("[chart unavailable]", styles["Normal"]))
    if trades:
        elements.append(Paragraph("Trades", styles["Heading2"]))
        cols = list(trades[0].keys())
        tdata = [cols] + [[str(r.get(c, "")) for c in cols] for r in trades[:100]]
        tt = Table(tdata, colWidths=[160 / len(cols) * mm] * len(cols))
        tt.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                ]
            )
        )
        elements.append(tt)
    doc.build(elements)
    raw = buf.getvalue()
    dq = dq or ExportQuality(0, 0.0, "unknown")
    fp = os.path.join(_ensure_dir(EXPORT_PDF_PATH), f"report_{run_id}.pdf")
    with open(fp, "wb") as f:
        f.write(raw)
    return ExportResult(
        run_id=run_id,
        format="pdf",
        data_quality=dq,
        data_hash=_data_hash(raw),
        file_path=fp,
        file_bytes=raw,
        created_at=_utc(),
        metadata={**meta, "title": title, "n_metrics": len(metrics)},
    )


def excel_report(
    metrics: dict,
    trades: list[dict],
    equity_curve: list[float] | None = None,
    drawdown: list[float] | None = None,
    dq: ExportQuality | None = None,
    seed: int | None = None,
) -> ExportResult:
    """Build multi-sheet Excel workbook (Summary, Trades, Equity, Drawdown, Quality)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    run_id = _run_id()
    meta = report_metadata(run_id, dq, seed)
    wb = Workbook()
    hfont, hfill = Font(bold=True, color="FFFFFF"), PatternFill(
        "solid", fgColor="2C3E50"
    )
    dq = dq or ExportQuality(0, 0.0, "unknown")

    def _hdr(ws, cols, row=1):
        for c, name in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=name)
            cell.font, cell.fill = hfont, hfill

    ws = wb.active
    ws.title = "Summary"
    _hdr(ws, ["Metric", "Value"])
    for i, (k, v) in enumerate(metrics.items(), 2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws2 = wb.create_sheet("Trades")
    if trades:
        cols = list(trades[0].keys())
        _hdr(ws2, cols)
        for i, r in enumerate(trades, 2):
            for c, name in enumerate(cols, 1):
                ws2.cell(row=i, column=c, value=r.get(name, ""))
    else:
        _hdr(ws2, ["symbol", "side", "qty", "price", "date"])
    for sheet_name, curve in [("Equity", equity_curve), ("Drawdown", drawdown)]:
        wsn = wb.create_sheet(sheet_name)
        _hdr(wsn, ["Step", sheet_name])
        for i, v in enumerate(curve or [], 2):
            wsn.cell(row=i, column=1, value=i - 1)
            wsn.cell(row=i, column=2, value=round(v, 6))
    ws5 = wb.create_sheet("Quality")
    _hdr(ws5, ["Field", "Value"])
    quality_rows = [
        ("system_version", meta["system_version"]),
        ("run_id", run_id),
        ("seed", meta["seed"]),
        ("created_at", meta["created_at"]),
        ("disclaimer", meta["disclaimer"]),
        ("n_observations", dq.n_observations),
        ("missing_pct", dq.missing_pct),
        ("source", dq.source),
        ("start_date", dq.start_date),
        ("end_date", dq.end_date),
    ]
    for i, (k, v) in enumerate(quality_rows, 2):
        ws5.cell(row=i, column=1, value=k)
        ws5.cell(row=i, column=2, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    fp = os.path.join(_ensure_dir(EXPORT_EXCEL_PATH), f"report_{run_id}.xlsx")
    with open(fp, "wb") as f:
        f.write(raw)
    return ExportResult(
        run_id=run_id,
        format="excel",
        data_quality=dq,
        data_hash=_data_hash(raw),
        file_path=fp,
        file_bytes=raw,
        created_at=_utc(),
        metadata={"sheets": wb.sheetnames},
    )
