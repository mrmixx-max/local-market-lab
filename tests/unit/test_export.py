"""Tests for export functions: PDF, Excel, CSV."""
from __future__ import annotations

import io

import numpy as np
import pytest

from packages.reports.export import (
    csv_equity,
    csv_scenario,
    csv_trades,
    excel_report,
    pdf_report,
)
from packages.domain.entities import ExportQuality


@pytest.fixture
def dq():
    return ExportQuality(n_observations=100, missing_pct=0.02, source="yahoo",
                       start_date="2024-01-01", end_date="2024-12-31")


class TestCsvExports:
    def test_csv_trades_basic(self, dq):
        trades = [
            {"symbol": "AAPL", "side": "buy", "qty": 10, "price": 150.0, "date": "2026-01-01"},
            {"symbol": "AAPL", "side": "sell", "qty": 5, "price": 155.0, "date": "2026-01-02"},
        ]
        result = csv_trades(trades, dq)
        assert result.format == "csv"
        assert result.run_id
        assert result.data_hash
        assert result.data_quality.source == "yahoo"
        assert result.metadata["rows"] == 2
        assert "AAPL" in open(result.file_path).read()

    def test_csv_trades_empty(self):
        result = csv_trades([])
        assert result.run_id
        assert result.metadata["rows"] == 0

    def test_csv_equity_basic(self, dq):
        curve = [100.0, 101.0, 99.5, 102.0]
        result = csv_equity(curve, dq=dq)
        assert result.format == "csv"
        assert result.metadata["rows"] == 4
        assert result.data_quality.n_observations == 100

    def test_csv_equity_with_dates(self, dq):
        curve = [100.0, 101.0]
        dates = ["2026-01-01", "2026-01-02"]
        result = csv_equity(curve, dates, dq)
        content = open(result.file_path).read()
        assert "2026-01-01" in content

    def test_csv_scenario(self, dq):
        runs = [{"run": 1, "terminal": 1.05}, {"run": 2, "terminal": 0.95}]
        result = csv_scenario(runs, dq)
        assert result.metadata["kind"] == "scenario"
        assert result.data_hash


class TestPdfExport:
    def test_pdf_report_basic(self, dq):
        metrics = {"sharpe": 1.2, "cagr": 0.08, "max_dd": -0.15}
        trades = [{"symbol": "TEST", "side": "buy", "qty": 10, "price": 100}]
        result = pdf_report("Test Report", metrics, trades, dq=dq)
        assert result.format == "pdf"
        assert result.run_id
        assert result.data_hash
        assert result.data_quality.source == "yahoo"

    def test_pdf_report_empty(self):
        result = pdf_report("Empty", {}, [])
        assert result.run_id
        assert result.data_hash

    def test_pdf_report_with_equity(self, dq):
        metrics = {"sharpe": 1.0}
        curve = [100.0 + i for i in range(50)]
        result = pdf_report("Equity", metrics, [], equity_curve=curve, dq=dq)
        assert result.metadata["n_metrics"] == 1


class TestExcelExport:
    def test_excel_report_basic(self, dq):
        from openpyxl import load_workbook
        metrics = {"sharpe": 1.2, "cagr": 0.08}
        trades = [{"symbol": "TEST", "side": "buy", "qty": 10, "price": 100}]
        result = excel_report(metrics, trades, dq=dq)
        assert result.format == "excel"
        assert result.run_id
        assert result.data_hash
        assert "Quality" in result.metadata["sheets"]

    def test_excel_report_sheets_content(self, dq):
        from openpyxl import load_workbook
        metrics = {"sharpe": 1.5, "volatility": 0.12}
        result = excel_report(metrics, [], dq=dq)
        wb = load_workbook(io.BytesIO(result.file_bytes or b""))
        assert "Summary" in wb.sheetnames
        assert "Trades" in wb.sheetnames
        assert "Equity" in wb.sheetnames
        assert "Drawdown" in wb.sheetnames
        assert "Quality" in wb.sheetnames

    def test_excel_report_with_curves(self, dq):
        from openpyxl import load_workbook
        curve = [100.0, 101.0, 102.0]
        dd = [0.0, 0.01, 0.02]
        result = excel_report({}, [], curve, dd, dq)
        wb = load_workbook(io.BytesIO(result.file_bytes or b""))
        ws = wb["Equity"]
        assert ws.cell(row=2, column=2).value == 100.0
        ws2 = wb["Drawdown"]
        assert ws2.cell(row=2, column=2).value == 0.0
