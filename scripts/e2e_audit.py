"""Full end-to-end release audit run — executed twice, results compared.

Covers: import -> quality -> hash -> features -> walk-forward -> purged CV
-> hyperparameter tuning -> model comparison (DM) -> explainability ->
backtest with costs -> stress tests -> rebalancing -> PDF/Excel/CSV export.

Writes a JSON result file; the runner executes it twice and diffs.
"""

from __future__ import annotations

import hashlib
import json
import os
import sys
import tempfile

SEED = 42
OUT = sys.argv[1] if len(sys.argv) > 1 else "e2e_result.json"
os.environ["LML_EXPORT_PDF_PATH"] = tempfile.mkdtemp(prefix="lml_pdf_")
os.environ["LML_EXPORT_EXCEL_PATH"] = tempfile.mkdtemp(prefix="lml_xls_")
os.environ["LML_EXPORT_CSV_PATH"] = tempfile.mkdtemp(prefix="lml_csv_")

R = {"steps": {}}


def step(name, fn):
    R["steps"][name] = fn()
    print(f"  ok: {name}")


# ---------- synthetic demo dataset (deterministic) ----------
import random

rng = random.Random(SEED)
DATES, PX = [], []
d = __import__("datetime").date(2020, 1, 1)
px = 100.0
while len(DATES) < 750:
    if d.weekday() < 5:
        px *= 1 + rng.gauss(0.0004, 0.012)
        DATES.append(d.isoformat())
        PX.append(round(px, 6))
    d += __import__("datetime").timedelta(days=1)

# ---------- data hash ----------
step(
    "data_hash",
    lambda: {
        "sha256": hashlib.sha256(
            json.dumps([DATES, PX], sort_keys=True).encode()
        ).hexdigest()[:16]
    },
)


# ---------- data quality ----------
def _quality():
    from packages.marketdata.series import series_quality
    from packages.domain.entities import PriceBar, PriceSeries

    s = PriceSeries("DEMO", "EUR", [PriceBar(dt, c) for dt, c in zip(DATES, PX)])
    q = series_quality(s)
    from packages.quality.checks import check_timestamps

    issues, dupes = check_timestamps(s)
    return {
        "points": q["points"],
        "dupes": dupes,
        "issues": len(issues),
        "nonmonotonic": sum(1 for i in issues if "non-monotonic" in i),
    }


step("quality", _quality)


# ---------- features ----------
def _features():
    import numpy as np

    rets = np.diff(np.log(PX))
    return {
        "n_returns": int(len(rets)),
        "mean": float(rets.mean()),
        "std": float(rets.std()),
    }


step("features", _features)


# ---------- validation ----------
def _wf():
    from packages.validation.walk_forward import walk_forward_backtest

    r = walk_forward_backtest(PX, lambda tr, te: [1.0] * len(te), seed=SEED)
    return r.summary()


step("walk_forward", _wf)


def _cv():
    from packages.validation.cv import time_series_cv

    r = time_series_cv(lambda tr, te: [1.0] * len(te), PX, seed=SEED)
    return r.summary()


step("purged_cv", _cv)


def _tune():
    from packages.validation.hyperparameter import hyperparameter_tune

    r = hyperparameter_tune(
        lambda tr, te, **p: [p.get("w", 1.0)] * len(te),
        PX,
        {"w": [0.5, 1.0, 1.5]},
        n_trials=3,
        seed=SEED,
        method="grid",
    )
    return r.summary()


step("hyperparameter", _tune)


# ---------- model comparison + explainability ----------
def _compare():
    import numpy as np
    from packages.explainability.comparison import WalkForwardResult, compare_models

    a = np.array(PX[300:])
    mk = lambda name, noise: [
        WalkForwardResult(
            window=w,
            train_start=0,
            train_end=w,
            test_start=w,
            test_end=w + 63,
            model_name=name,
            mse=float((((a[w : w + 63] * (1 + noise)) - a[w : w + 63]) ** 2).mean()),
            mae=float((np.abs(a[w : w + 63] * (1 + noise) - a[w : w + 63])).mean()),
        )
        for w in range(0, 400, 100)
    ]
    cm = compare_models(mk("model-a", 0.001), mk("model-b", 0.004)).to_dict()
    return {"dm": cm["diebold_mariano"]}


step("model_comparison", _compare)


def _explain():
    import numpy as np
    from packages.explainability.importance import (
        permutation_importance,
        shapley_approx,
    )

    rng2 = np.random.default_rng(SEED)
    X = rng2.standard_normal((200, 4))
    y = X[:, 1] * 3 + rng2.standard_normal(200) * 0.2
    pi = permutation_importance(
        lambda X_: X_[:, 1] * 3, X, y, n_repeats=5, seed=SEED
    ).to_dict()
    sh = shapley_approx(lambda X_: X_[:, 1] * 3, X, X[0], n_samples=30, seed=SEED)
    return {
        "top_feature": max(pi["feature_importance"], key=lambda f: f["importance"])[
            "feature"
        ],
        "shap_approx_flag": sh["approximation"],
    }


step("explainability", _explain)


# ---------- backtest with costs ----------
def _backtest():
    from packages.backtest.engine import (
        Assumptions,
        BuyAndHold,
        PeriodicRebalance,
        run_backtest,
    )

    prices = {"DEMO": PX}
    bh = run_backtest(prices, BuyAndHold(), Assumptions(fees_bps=10, slippage_bps=5))
    pr = run_backtest(
        prices, PeriodicRebalance(63), Assumptions(fees_bps=10, slippage_bps=5)
    )
    return {
        "bh_final": round(bh["curve"][-1], 6),
        "pr_final": round(pr["curve"][-1], 6),
        "bh_trades": bh["trades"],
        "pr_trades": pr["trades"],
        "costs_less_than_free": pr["curve"][-1] < bh["curve"][-1],
    }


step("backtest_costs", _backtest)


# ---------- stress + rebalancing ----------
def _stress():
    from packages.scenarios.stress import (
        HISTORICAL_CRISES,
        HYPOTHETICAL_SCENARIOS,
        monte_carlo_fat_tail,
        run_historical_stress,
        run_hypothetical_stress,
    )

    P = {"IWDA": 0.6, "AGGH": 0.4}
    out = {}
    for s in HISTORICAL_CRISES:
        out[s] = run_historical_stress(s, P, seed=SEED).metrics["max_drawdown"]
    for s in HYPOTHETICAL_SCENARIOS:
        out[s] = run_hypothetical_stress(s, P, seed=SEED).metrics["max_drawdown"]
    out["mc_p01"] = monte_carlo_fat_tail(P, runs=500, seed=SEED)["p01"]
    return out


step("stress", _stress)


def _rebal():
    from packages.portfolio.rebalancing import suggest_rebalance

    r = suggest_rebalance(
        {"IWDA": 0.72, "AGGH": 0.28},
        {"IWDA": 0.6, "AGGH": 0.4},
        transaction_cost_bps=10,
    )
    return {
        "proposals": [{"symbol": p.symbol, "action": p.action} for p in r.proposals],
        "total_cost": r.total_estimated_cost,
    }


step("rebalancing", _rebal)


# ---------- exports ----------
def _exports():
    from packages.reports.export import csv_equity, excel_report, pdf_report
    from packages.domain.entities import ExportQuality

    dq = ExportQuality(len(PX), 0.0, "synthetic", DATES[0], DATES[-1])
    curve = [round(p / PX[0] * 100, 6) for p in PX]
    c = csv_equity(curve, DATES, dq=dq, seed=SEED)
    x = excel_report(
        {"cagr_pct": 7.1}, [{"symbol": "DEMO", "qty": 10}], curve, dq=dq, seed=SEED
    )
    p = pdf_report("E2E Audit", {"cagr_pct": 7.1}, [], dq=dq, seed=SEED)

    # verify each file is readable and carries metadata
    head = open(c.file_path, encoding="utf-8").readline()
    assert "# system_version" in head, "CSV metadata header missing"
    from openpyxl import load_workbook

    wb = load_workbook(x.file_path)
    vals = {
        wb["Quality"]
        .cell(row=i, column=1)
        .value: wb["Quality"]
        .cell(row=i, column=2)
        .value
        for i in range(1, 16)
    }
    assert vals.get("system_version") and vals.get("seed") == SEED
    assert open(p.file_path, "rb").read(4) == b"%PDF"
    return {
        "csv_hash": c.data_hash,
        "xlsx_hash": x.data_hash,
        "pdf_hash": p.data_hash,
        "csv_meta_ok": True,
        "xlsx_meta_ok": True,
        "pdf_magic_ok": True,
    }


step("exports", _exports)

R["run_id_seed"] = SEED
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(R, f, indent=1, sort_keys=True)
print(f"E2E COMPLETE -> {OUT}")
